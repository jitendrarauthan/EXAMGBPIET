"""
PDF & Excel processing utilities for the Result Asterisk Portal.

- Parses TC (Tabulation Chart) PDFs: multi-student-per-page tabular dump.
- Parses GS (Provisional Grade Sheet) PDFs: one student per page.
- Parses SEM_X excel sheets: detects highlighted (yellow / blue) cells per
  (roll_no, subject_code) which represent BACK / failed subjects.
- Parses TC_X / GS_X excel sheets to extract structured per-student records
  directly (so admin can upload only an Excel file and we still generate all
  semester TC*/GS* PDFs).
- Generates new TC*/GS* PDFs in the same visual format with " *" appended
  to subject names where a back exists. GS pages also embed a Code-128
  barcode (encoded with the university roll number).
"""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pdfplumber
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

ROMAN = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII"]
ROMAN_TO_INT = {r: i + 1 for i, r in enumerate(ROMAN)}

# ---------------------------------------------------------------------------
# Excel parsing — SEM_X sheet -> highlighted (back) subjects per student
# ---------------------------------------------------------------------------

# default openpyxl "no fill" tokens
_NO_FILL = {"00000000", "FFFFFFFF", "0", None, ""}


def _is_highlighted(cell) -> bool:
    f = getattr(cell, "fill", None)
    if not f or not getattr(f, "patternType", None):
        return False
    if f.patternType != "solid":
        return False
    fg = f.fgColor
    if not fg:
        return False
    rgb = fg.rgb if fg.type == "rgb" else None
    if rgb in _NO_FILL:
        return False
    return rgb is not None  # any non-default solid fill counts


def parse_sem_excel(file_bytes: bytes) -> Dict[str, Dict[str, Dict[str, bool]]]:
    """Return {sem_roman: {roll_no: {subject_code: True/False(=back)}}}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out: Dict[str, Dict[str, Dict[str, bool]]] = {}
    for sname in wb.sheetnames:
        m = re.match(r"^SEM[_ ]?(I{1,3}|IV|V|VI{1,3}|VIII)$", sname.strip(), re.I)
        if not m:
            continue
        sem = m.group(1).upper()
        ws = wb[sname]
        out[sem] = _parse_back_map_from_ws(ws)
    return out


def _parse_back_map_from_ws(ws) -> Dict[str, Dict[str, bool]]:
    """Shared worksheet → {roll: {subject_code: True}} parser used by both
    the standard ``SEM_<sem>`` and the M.Tech ``SEM_<branch>`` workflows.

    Auto-detects the header row holding subject codes (M.Tech branch sheets
    place them on row 5 instead of row 1) and the column carrying the
    University Roll No.
    """
    # 1. Locate the header row that lists the subject codes. We look for a
    # row whose populated cells (after a leading "SC"/"M.M."/"SN" anchor)
    # include strings that LOOK like course codes (3+ alphabetic chars then a
    # space or dash and digits).
    header_row = 1
    for r in range(1, min(15, ws.max_row + 1)):
        codes_in_row = [
            ws.cell(r, c).value for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value
            and re.match(r"^[A-Z]{2,5}\s*-?\s*\d{2,4}$",
                          str(ws.cell(r, c).value).strip(), re.I)
        ]
        if len(codes_in_row) >= 2:
            header_row = r
            break

    # 2. Find the column where subject codes start
    subj_start_col = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and re.match(r"^[A-Z]{2,5}\s*-?\s*\d{2,4}$", str(v).strip(), re.I):
            subj_start_col = c
            break

    # 3. Build subject_code → list of column indices it spans
    subj_cols: List[Tuple[str, List[int]]] = []
    last_code: Optional[str] = None
    for c in range(subj_start_col, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip():
            txt = str(v).strip()
            if re.match(r"^[A-Z]{2,5}\s*-?\s*\d{2,4}$", txt, re.I):
                subj_cols.append((txt, [c]))
                last_code = txt
            elif subj_cols and last_code:
                subj_cols[-1][1].append(c)
        else:
            if subj_cols and last_code:
                subj_cols[-1][1].append(c)

    # 4. Find the column carrying the University Roll No. on the data rows.
    # In B.Tech SEM_X sheets the roll sits in col 2; in M.Tech branch sheets
    # also col 2 (after a serial-number col 1). We just probe rows below the
    # header to find a column with mostly numeric values >= 8 chars long.
    roll_col = 2
    sample_cols = [2, 3, 4]
    for cand in sample_cols:
        n_numeric = 0
        for r in range(header_row + 1, min(header_row + 12, ws.max_row + 1)):
            v = ws.cell(r, cand).value
            if v and any(ch.isdigit() for ch in str(v)) and len(str(v).strip()) >= 8:
                n_numeric += 1
        if n_numeric >= 2:
            roll_col = cand
            break

    # 5. Iterate student rows
    sem_map: Dict[str, Dict[str, bool]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        roll = ws.cell(r, roll_col).value
        if not roll:
            continue
        roll_s = str(roll).strip()
        if not any(ch.isdigit() for ch in roll_s):
            continue
        student_back: Dict[str, bool] = {}
        for code, cols in subj_cols:
            back = any(_is_highlighted(ws.cell(r, c)) for c in cols)
            if back:
                student_back[code] = True
        if student_back:
            sem_map[roll_s] = student_back
    return sem_map


def inspect_workbook_sheets(file_bytes: bytes) -> Dict[str, list]:
    """Return classification of an uploaded workbook's sheets — used by the
    M.Tech upload UI so the admin can pick which SEM_<branch>, TC_<branch>
    and GS_<branch> sheets to use.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = {"sem": [], "tc": [], "gs": [], "all": list(wb.sheetnames)}
    for sname in wb.sheetnames:
        s = sname.strip()
        if re.match(r"^SEM[_ ]", s, re.I):
            out["sem"].append(s)
        elif re.match(r"^TC[_ ]", s, re.I):
            out["tc"].append(s)
        elif re.match(r"^GS[_ ]", s, re.I):
            out["gs"].append(s)
    return out


def parse_sem_excel_sheet(file_bytes: bytes, sheet_name: str) -> Dict[str, bool]:
    """Parse a SINGLE named SEM_<X> sheet into a {roll: {subject: True}} back
    paper map (the M.Tech workflow uses ``SEM_<branch>`` sheet names where
    ``<branch>`` is e.g. CSE / Production / Thermal / ECE / BT etc.).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
    return _parse_back_map_from_ws(wb[sheet_name])


def parse_tc_or_gs_named_sheet(file_bytes: bytes, sheet_name: str, kind: str) -> List[dict]:
    """Parse a NAMED ``TC_*`` / ``GS_*`` sheet (e.g. ``TC_CSE`` for the
    M.Tech workflow) into per-student records using the same column logic as
    the standard parsers."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
    ws = wb[sheet_name]
    # M.Tech workbooks ship the data as repeating per-student blocks anchored
    # at a column-1 "M.M." marker. Detect that layout up front and dispatch
    # to the dedicated block parser; otherwise fall back to the standard
    # AIML-style sheet parsers.
    if _looks_like_mtech_block_sheet(ws):
        return parse_mtech_block_sheet(ws, kind)
    if kind.lower() == "tc":
        return parse_tc_excel_sheet(ws)
    if kind.lower() == "gs":
        return parse_gs_excel_sheet(ws)
    raise ValueError(f"Unknown sheet kind: {kind}")


def _looks_like_mtech_block_sheet(ws) -> bool:
    """Heuristic: M.Tech Marksheets / Grade Sheets sheets contain repeating
    'M.M.' anchor rows in column 1 followed by per-student blocks. Scan the
    first ~50 rows for any such marker."""
    for r in range(1, min(50, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and str(v).strip().upper().startswith("M.M."):
            return True
    return False


def parse_mtech_block_sheet(ws, kind: str) -> List[dict]:
    """Parse the per-student block format used in M.Tech Marksheets and
    Grade Sheets sheets.

    Both layouts begin each student with a ``M.M.`` marker row in column 1.
    The Marksheets layout is rich (carries external / sessional / total
    marks); the Grade Sheets layout is leaner. We detect which one we're
    looking at by inspecting where the student name lands relative to the
    "Name:" label.
    """
    # Find every block start row
    starts: List[int] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip().upper().startswith("M.M."):
            starts.append(r)
    if not starts:
        return []
    starts.append(ws.max_row + 1)  # sentinel

    out: List[dict] = []
    for i in range(len(starts) - 1):
        r0, r1 = starts[i], starts[i + 1]
        rec = _parse_mtech_one_block(ws, r0, r1, kind)
        if rec and rec.get("roll_no"):
            out.append(rec)
    return out


def _scan_label_row(ws, r: int, rec: dict) -> None:
    """Scan one row for known labels and stash their adjacent values into the
    student record. Used by the M.Tech block parser where the same logical
    fields (Result / SGPA / CGPA / Earned Credits / Remark) are placed in
    different column slots in different sheet layouts."""
    LABEL_TO_KEY = [
        ("RESULT", "result"),
        ("REMARK", "remark"),
        ("SGPA", "sgpa"),
        ("CGPA", "cgpa"),
        ("EARNED CREDIT", "earned_credits"),
        ("CUML", "cuml_earned_credits"),
    ]
    LABELS = [lbl for lbl, _ in LABEL_TO_KEY]
    max_c = ws.max_column

    def _is_another_label(text: str) -> bool:
        s = text.strip().upper().rstrip(":").strip()
        return any(s.startswith(lbl) for lbl in LABELS)

    for c in range(1, max_c + 1):
        v = ws.cell(r, c).value
        if v is None:
            continue
        s = str(v).strip().upper().rstrip(":").strip()
        matched = next((k for lbl, k in LABEL_TO_KEY if s.startswith(lbl)), None)
        if not matched:
            continue
        # Look at the next ≤3 cells; pick the first non-empty value, but stop
        # if we encounter another label (means our value cell is genuinely
        # blank — e.g., empty Remark).
        for cc in range(c + 1, min(c + 4, max_c + 1)):
            nv = ws.cell(r, cc).value
            if nv is None:
                continue
            txt = str(nv).strip()
            if txt == "":
                continue
            if _is_another_label(txt):
                break
            if not rec.get(matched):
                rec[matched] = txt
            break


def _parse_mtech_one_block(ws, r0: int, r1: int, kind: str) -> Optional[dict]:
    """Parse a SINGLE per-student block bounded by [r0, r1) rows."""
    rec: dict = {
        "name": "", "father_name": "", "roll_no": "", "enroll_no": "",
        "subjects": [],
        "sgpa": "", "cgpa": "", "result": "", "remark": "",
        "earned_credits": "", "cuml_earned_credits": "",
        "program": "", "branch": "", "semester": "", "exam_session": "",
    }
    subj_header_row = None

    def _g(r: int, c: int) -> str:
        v = ws.cell(r, c).value
        return "" if v is None else str(v).strip()

    for r in range(r0, r1):
        c1 = _g(r, 1)
        if c1.lower().startswith("name:"):
            # Marksheets layout puts name in col 2 (after "Name:" in col 1).
            # Grade Sheets layout uses col 3 with col 1 = "Name:" and col 5
            # holding the "University Roll No" label, so try multiple cols.
            for c in (2, 3):
                if _g(r, c):
                    rec["name"] = _g(r, c)
                    break
            # Father / roll / enroll on the SAME row in Marksheets layout
            for col, key in ((5, "father_name"), (9, "roll_no"), (13, "enroll_no")):
                v = _g(r, col)
                if v:
                    rec[key] = v
            # Roll usually sits at col 9 in both layouts
            for col in (9, 6, 11):
                if not rec["roll_no"] and _g(r, col):
                    if any(ch.isdigit() for ch in _g(r, col)):
                        rec["roll_no"] = _g(r, col)
                        break
        elif c1.lower().startswith("father"):
            # Grade Sheets layout: father / enroll on a separate row
            for c in (3, 2):
                if _g(r, c):
                    rec["father_name"] = _g(r, c)
                    break
            for col in (9, 11):
                v = _g(r, col)
                if v and "/" in v or (v and v.upper().startswith("UTU")):
                    rec["enroll_no"] = v
                    break
        elif c1.lower().startswith("subject code"):
            subj_header_row = r
        elif c1.lower().startswith(("total credits", "total credit", "total ")):
            # Skip — totals row
            continue
        elif c1.lower().startswith("result"):
            # Find values by scanning labelled cells across this row.
            # Different M.Tech sheet layouts place the values in different
            # columns, so we look for each label and pick up the next
            # non-empty cell to its right.
            _scan_label_row(ws, r, rec)
        elif c1.lower().startswith("earned credit"):
            # GS layout: Earned Credits + SGPA + CGPA share one row.
            _scan_label_row(ws, r, rec)
        else:
            # Subject line: only consider rows AFTER the header row, with a
            # plausible subject-code in col 1.
            if subj_header_row and r > subj_header_row and c1 and not c1.lower().startswith("name"):
                # Marksheets columns: 1=code, 2=name, 6=credits, 7=external,
                # 9=sessional, 11=total, 13=grade, 14=gp.
                # Grade Sheets columns: 1=code, 3=name, 8=credits, 9=grade,
                # 10=gp.
                code = c1
                # name column varies — pick the first non-empty in (2, 3)
                name = _g(r, 2) or _g(r, 3)
                # detect layout: if col 7 has a value matching N/N format -> Marksheets
                external = _g(r, 7)
                sessional = _g(r, 9)
                total = _g(r, 11)
                grade = _g(r, 13)
                gp = _g(r, 14)
                credits = _g(r, 6)
                if not external or "/" not in external:
                    # Grade Sheets layout
                    credits = _g(r, 8) or credits
                    grade = _g(r, 9) or grade
                    gp = _g(r, 10) or gp
                    external = ""
                    sessional = ""
                    total = ""
                if not name:
                    continue
                # Treat Excel error placeholders (#N/A, #VALUE! etc.) as empty.
                def _clean(x: str) -> str:
                    return "" if x.startswith("#") else x
                rec["subjects"].append({
                    "code": code,
                    "name": name,
                    "credits": _clean(credits),
                    "external": _clean(external),
                    "sessional": _clean(sessional),
                    "total": _clean(total),
                    "grade": _clean(grade),
                    "grade_points": _clean(gp),
                })

    # Default ``result`` if missing — derive from grades.
    if not rec["result"] and rec["subjects"]:
        if any((s.get("grade") or "").upper() in ("F", "AB", "DT", "UFM")
               for s in rec["subjects"]):
            rec["result"] = "FAIL"
        else:
            rec["result"] = "PASS"

    # Round numeric SGPA / CGPA / Earned-Credit strings to 2 decimals so the
    # GS / TC don't display the full IEEE-754 expansion (6.6363636363...).
    for k in ("sgpa", "cgpa"):
        v = rec.get(k, "")
        if v:
            try:
                rec[k] = f"{float(v):.2f}"
            except (TypeError, ValueError):
                pass
    for k in ("earned_credits", "cuml_earned_credits"):
        v = rec.get(k, "")
        if v:
            try:
                f = float(v)
                rec[k] = str(int(f)) if f == int(f) else f"{f:.2f}"
            except (TypeError, ValueError):
                pass

    return rec if rec.get("roll_no") else None


# ---------------------------------------------------------------------------
# Excel parsing — TC_X and GS_X sheets -> per-student records
# ---------------------------------------------------------------------------

_SEM_RE = re.compile(r"^(TC|GS)[_ ]?(I{1,3}|IV|V|VI{1,3}|VIII)$", re.I)


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def _is_grade(g: str) -> bool:
    return bool(g) and bool(re.match(r"^([A-Z][+\-]?|Excellent|Ab|F|P|D|E)$", g))


_HEADER_BLOCK_RE = re.compile(
    r"(?P<program>(Bachelor|Master).+?)\n+(?P<branch>.+?)\n+(?P<sem>[IVX]+)\s+Semester\s+(?P<session>[A-Za-z]+\s+\d{4})",
    re.S,
)


def _extract_sheet_meta(ws) -> Dict[str, str]:
    """Extract program / branch / semester / session from the top header rows
    of a TC_X or GS_X sheet (the first cell that mentions 'Bachelor' or
    'Master')."""
    meta: Dict[str, str] = {"program": "", "branch": "", "semester": "",
                              "exam_session": ""}
    blob_parts: List[str] = []
    for r in range(1, 8):
        for c in range(1, 8):
            v = ws.cell(r, c).value
            if v:
                blob_parts.append(str(v))
    blob = "\n".join(blob_parts)
    blob = re.sub(r"Tabulation Chart for\s*", "", blob)
    blob = re.sub(r"GRADE SHEET\s*", "", blob)
    m = _HEADER_BLOCK_RE.search(blob)
    if m:
        meta["program"] = m.group("program").strip()
        meta["branch"] = re.sub(r"\s{2,}", " ", m.group("branch")).strip()
        meta["semester"] = m.group("sem").strip()
        meta["exam_session"] = m.group("session").strip()
    return meta


def parse_tc_excel_sheet(ws) -> List[dict]:
    """Parse a TC_X sheet -> list of student records.

    Block boundaries: rows starting with 'MM' in column A (case-insensitive).
    """
    records: List[dict] = []
    sheet_meta = _extract_sheet_meta(ws)
    rows = ws.max_row
    block_starts: List[int] = []
    for r in range(1, rows + 1):
        v = _norm(ws.cell(r, 1).value)
        if v.upper() == "MM":
            if _norm(ws.cell(r, 3).value).upper() == "SC":
                block_starts.append(r)
    block_starts.append(rows + 1)

    for i in range(len(block_starts) - 1):
        r0, r1 = block_starts[i], block_starts[i + 1]
        rec: dict = {"subjects": [], **sheet_meta}
        hdr_row = r0 + 1
        if hdr_row >= r1:
            continue
        rec["name"] = _norm(ws.cell(hdr_row, 2).value)
        rec["father_name"] = _norm(ws.cell(hdr_row, 5).value)
        rec["roll_no"] = _norm(ws.cell(hdr_row, 9).value)
        rec["enroll_no"] = _norm(ws.cell(hdr_row, 13).value)
        if not rec["roll_no"]:
            continue
        for rr in range(hdr_row + 2, r1):
            code = _norm(ws.cell(rr, 1).value)
            low = code.lower()
            # ---- Result / summary row ----
            if low.startswith("result"):
                rec["result"] = _norm(ws.cell(rr, 2).value)
                rec["remark"] = _norm(ws.cell(rr, 5).value)
                rec["sgpa"] = _norm(ws.cell(rr, 8).value)
                rec["cgpa"] = _norm(ws.cell(rr, 10).value)
                rec["earned_credits"] = _norm(ws.cell(rr, 12).value)
                rec["cuml_earned_credits"] = _norm(ws.cell(rr, 14).value)
                continue
            if not code or low.startswith("subject") or low.startswith("total"):
                continue
            if not re.match(r"^[A-Z]{2,4}\s?\d{2,4}$", code):
                continue
            name = _norm(ws.cell(rr, 2).value)
            credits = _norm(ws.cell(rr, 6).value)
            ext = _norm(ws.cell(rr, 7).value)
            ses = _norm(ws.cell(rr, 9).value)
            tot = _norm(ws.cell(rr, 11).value)
            grade = _norm(ws.cell(rr, 13).value)
            gp = _norm(ws.cell(rr, 14).value)
            rec["subjects"].append({
                "code": code,
                "name": name,
                "credits": credits,
                "external": ext,
                "sessional": ses,
                "total": tot,
                "grade": grade,
                "grade_points": gp,
                "back": False,
            })
        for k in ("sgpa", "cgpa"):
            v = rec.get(k, "")
            try:
                rec[k] = f"{float(v):.2f}"
            except Exception:
                pass
        if rec["subjects"]:
            records.append(rec)
    return records


def parse_gs_excel_sheet(ws) -> List[dict]:
    """Parse a GS_X sheet -> list of student records.

    Block boundaries: rows where col A starts with 'Name:' AND col F contains 'University Roll'.
    """
    records: List[dict] = []
    sheet_meta = _extract_sheet_meta(ws)
    rows = ws.max_row
    block_starts: List[int] = []
    for r in range(1, rows + 1):
        v = _norm(ws.cell(r, 1).value)
        f = _norm(ws.cell(r, 6).value)
        if v.startswith("Name:") and f.startswith("University Roll"):
            block_starts.append(r)
    block_starts.append(rows + 1)

    for i in range(len(block_starts) - 1):
        r0, r1 = block_starts[i], block_starts[i + 1]
        rec: dict = {"subjects": [], **sheet_meta}
        rec["name"] = _norm(ws.cell(r0, 3).value)
        rec["roll_no"] = _norm(ws.cell(r0, 9).value)
        rec["father_name"] = _norm(ws.cell(r0 + 1, 3).value)
        rec["enroll_no"] = _norm(ws.cell(r0 + 1, 9).value)
        if not rec["roll_no"]:
            continue
        # Subjects start after the "Subject Code" header row
        for rr in range(r0 + 3, r1):
            code = _norm(ws.cell(rr, 1).value)
            name = _norm(ws.cell(rr, 3).value)
            credits = _norm(ws.cell(rr, 8).value)
            grade = _norm(ws.cell(rr, 9).value)
            gp = _norm(ws.cell(rr, 10).value)
            low = code.lower()
            if not code:
                continue
            if low.startswith("total"):
                continue
            if low.startswith("earned"):
                # Earned Credits / SGPA / CGPA row
                rec["earned_credits"] = _norm(ws.cell(rr, 3).value)
                rec["sgpa"] = _norm(ws.cell(rr, 6).value)
                rec["cgpa"] = _norm(ws.cell(rr, 9).value)
                continue
            if low.startswith("result"):
                rec["result"] = _norm(ws.cell(rr, 2).value)
                rec["remark"] = _norm(ws.cell(rr, 6).value)
                continue
            if re.match(r"^[A-Z]{2,4}\s?\d{2,4}$", code):
                rec["subjects"].append({
                    "code": code,
                    "name": name,
                    "credits": credits,
                    "grade": grade,
                    "grade_points": gp,
                    "back": False,
                })
        for k in ("sgpa", "cgpa"):
            v = rec.get(k, "")
            try:
                rec[k] = f"{float(v):.2f}"
            except Exception:
                pass
        if rec["subjects"]:
            records.append(rec)
    return records


def parse_tc_gs_excel(file_bytes: bytes) -> Dict[str, Dict[str, List[dict]]]:
    """Parse all TC_X / GS_X sheets. Returns {'TC': {sem: records}, 'GS': {sem: records}}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out: Dict[str, Dict[str, List[dict]]] = {"TC": {}, "GS": {}}
    for sname in wb.sheetnames:
        m = _SEM_RE.match(sname.strip())
        if not m:
            continue
        kind = m.group(1).upper()
        sem = m.group(2).upper()
        ws = wb[sname]
        if kind == "TC":
            out["TC"][sem] = parse_tc_excel_sheet(ws)
        else:
            out["GS"][sem] = parse_gs_excel_sheet(ws)
    return out


# ---------------------------------------------------------------------------
# PDF parsing — TC (multi-student per page)
# ---------------------------------------------------------------------------

_HEADER_PATTERNS = [
    "GOVIND BALLABH PANT INSTITUTE",
    "PAURI GARHWAL",
    "(AN AUTONOMOUS",
    "(AFFILIATED TO",
]


def _detect_program_branch_sem(text: str) -> Tuple[str, str, str]:
    """Return (program, branch, semester_roman) from a TC/GS page header."""
    program = ""
    branch = ""
    sem = ""
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "Tabulation Chart for" in line:
            program = line.replace("Tabulation Chart for", "").strip()
        elif line.startswith("Bachelor of") or line.startswith("Master of"):
            program = line
        m = re.match(r"^([IVX]+)\s+Semester\s+([A-Za-z]+\s+\d{4})", line)
        if m:
            sem = m.group(1).upper()
            continue
        # Branch is usually the line right after program if it's not the sem line
        if program and not sem and line != program and "Semester" not in line:
            if not any(h in line for h in _HEADER_PATTERNS):
                if not branch:
                    branch = line
    return program, branch, sem


def parse_tc_pdf(file_bytes: bytes) -> List[dict]:
    """Parse Tabulation Chart PDF -> list of student records."""
    students: List[dict] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        program, branch, sem = "", "", ""
        all_text = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            p, b, s = _detect_program_branch_sem(text)
            program = p or program
            branch = b or branch
            sem = s or sem
            all_text.append(text)
        full = "\n".join(all_text)

        # Each student block starts with "MM N SC SN C ..." marker line.
        blocks = re.split(r"\n(?=MM\s+\d+\s+SC\s+SN\s+C\b)", full)
        for blk in blocks:
            if "University Roll No" not in blk:
                continue
            rec = _parse_tc_block(blk)
            if rec:
                rec["program"] = program
                rec["branch"] = branch
                rec["semester"] = sem
                students.append(rec)
    return students


def _parse_tc_block(block: str) -> Optional[dict]:
    rec: dict = {"subjects": []}

    # Roll & enrollment
    roll = re.search(r"University Roll No:?\s*(\S+)", block)
    enr = re.search(r"University Enrol(?:lment|\.)?\s*No:?\s*(\S+)", block)
    if not roll:
        return None
    rec["roll_no"] = roll.group(1).strip()
    rec["enroll_no"] = enr.group(1).strip() if enr else ""

    # The "Name:" line in the TC PDF collapses Name and Father's Name on a
    # single line: "Name: <Student> <Father> University Roll No: ..."
    # We extract the trailing student name from the MM line if present.
    mm = re.search(r"^MM\s+\d+\s+SC\s+SN\s+C\s+(.+)$", block, re.M)
    student_in_mm = ""
    if mm:
        candidate = mm.group(1).strip()
        if candidate and not re.match(r"^[IVX]+$", candidate):
            student_in_mm = candidate
    name_line = re.search(
        r"Name:\s*(.+?)\s+University Roll No:?\s*\S+",
        block.replace("\n", " "),
    )
    name_block = name_line.group(1).strip() if name_line else ""
    if student_in_mm and student_in_mm in name_block:
        father = name_block.replace(student_in_mm, "", 1).strip()
        rec["name"] = student_in_mm
        rec["father_name"] = father
    else:
        # No reliable split — store full blob in name; admin can correct via GS upload
        parts = name_block.split()
        if len(parts) >= 2:
            mid = len(parts) // 2
            rec["name"] = " ".join(parts[:mid])
            rec["father_name"] = " ".join(parts[mid:])
        else:
            rec["name"] = name_block
            rec["father_name"] = ""

    # Subjects: lines like "AHT 002 Engineering Chemistry 4 45/100 37/50 82/150 E+ 5.5"
    subj_re = re.compile(
        r"^([A-Z]{2,4}\s?\d{2,4})\s+(.+?)\s+(\d+)\s+([\d\-]+/\d+|-)\s+(\d+/\d+)\s+(\d+/\d+)\s+([A-Z][+\-]?|Excellent|F)\s*([\d.]+)?\s*$"
    )
    for line in block.splitlines():
        line = line.strip()
        m = subj_re.match(line)
        if m:
            rec["subjects"].append({
                "code": m.group(1).strip(),
                "name": m.group(2).strip(),
                "credits": m.group(3),
                "external": m.group(4),
                "sessional": m.group(5),
                "total": m.group(6),
                "grade": m.group(7),
                "grade_points": m.group(8) or "",
                "back": False,
            })

    # Result / SGPA / CGPA
    flat = block.replace("\n", " ")
    sg = re.search(r"SGPA:\s*([\d.]+)", flat)
    cg = re.search(r"CGPA:\s*([\d.]+)", flat)
    res = re.search(r"Result:\s*([A-Z ]+?)\s+(?:Remark|SGPA)", flat)
    rem = re.search(r"Remark:\s*(.*?)\s*(?=SGPA|Cuml|Earned|MM\s+\d+|$)", flat)
    rec["sgpa"] = sg.group(1) if sg else ""
    rec["cgpa"] = cg.group(1) if cg else ""
    rec["result"] = res.group(1).strip() if res else ""
    raw_rem = (rem.group(1).strip() if rem else "")
    rec["remark"] = raw_rem[:200]
    return rec if rec.get("subjects") else None


# ---------------------------------------------------------------------------
# PDF parsing — GS (one student per page)
# ---------------------------------------------------------------------------


def parse_gs_pdf(file_bytes: bytes) -> List[dict]:
    students: List[dict] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        program, branch, sem = "", "", ""
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                ls = line.strip()
                if ls.startswith("Bachelor of") or ls.startswith("Master of"):
                    program = ls
                m = re.match(r"^([IVX]+)\s+Semester\s+([A-Za-z]+\s+\d{4})", ls)
                if m:
                    sem = m.group(1).upper()
            # branch line is between program and semester
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            for i, l in enumerate(lines):
                if l == program and i + 1 < len(lines):
                    nxt = lines[i + 1]
                    if "Semester" not in nxt and not any(h in nxt for h in _HEADER_PATTERNS):
                        branch = nxt
                        break

            rec = _parse_gs_page(text)
            if rec:
                rec["program"] = program
                rec["branch"] = branch
                rec["semester"] = sem
                students.append(rec)
    return students


def _parse_gs_page(text: str) -> Optional[dict]:
    rec: dict = {"subjects": []}
    flat = text.replace("\n", " ")

    m = re.search(
        r"Name:\s*(.+?)\s+University Roll No\.?:\s*(\S+).*?Father'?s?\s*Name:\s*(.+?)\s+University Enrollment No\.?:\s*(\S+)",
        flat,
    )
    if m:
        rec["name"] = m.group(1).strip()
        rec["roll_no"] = m.group(2).strip()
        rec["father_name"] = m.group(3).strip()
        rec["enroll_no"] = m.group(4).strip()
    else:
        roll = re.search(r"University Roll No\.?:\s*(\S+)", flat)
        if not roll:
            return None
        rec["roll_no"] = roll.group(1).strip()
        rec["name"] = ""
        rec["father_name"] = ""
        rec["enroll_no"] = ""

    sl = re.search(r"SL\.?\s*NO\.?:\s*(\S+)", flat)
    rec["sl_no"] = sl.group(1) if sl else ""

    subj_re = re.compile(
        r"^([A-Z]{2,4}\s?\d{2,4})\s+(.+?)\s+(\d+)\s+([A-Z][+\-]?|Excellent|F)\s*([\d.]+)?\s*$"
    )
    for line in text.splitlines():
        ls = line.strip()
        m2 = subj_re.match(ls)
        if m2:
            rec["subjects"].append({
                "code": m2.group(1).strip(),
                "name": m2.group(2).strip(),
                "credits": m2.group(3),
                "grade": m2.group(4),
                "grade_points": m2.group(5) or "",
                "back": False,
            })

    sg = re.search(r"SGPA:\s*([\d.]+)", flat)
    cg = re.search(r"CGPA:\s*([\d.]+)", flat)
    ec = re.search(r"Earned Credits:\s*(\d+)", flat)
    res = re.search(r"Result:\s*([A-Z ]+?)\s+Remark:", flat)
    rem = re.search(r"Remark:\s*(.+?)(?:Prepared by|$)", flat)
    rec["sgpa"] = sg.group(1) if sg else ""
    rec["cgpa"] = cg.group(1) if cg else ""
    rec["earned_credits"] = ec.group(1) if ec else ""
    rec["result"] = res.group(1).strip() if res else ""
    rec["remark"] = (rem.group(1).strip() if rem else "")[:200]
    return rec if rec.get("subjects") else None


# ---------------------------------------------------------------------------
# Apply asterisk markers
# ---------------------------------------------------------------------------


def apply_back_markers(records: List[dict], back_map_for_sem: Dict[str, Dict[str, bool]]) -> Tuple[int, int]:
    """Mutate records: append ' *' to subject name when SEM excel marks it as
    back AND the student has cleared the paper (i.e. grade is not F / Ab /
    Dt). The '*' therefore means *subject cleared after a back paper*.

    Returns (markers_applied, students_matched).
    """
    n = 0
    matched = 0
    record_rolls = {r.get("roll_no", "") for r in records}
    skip_grades = {"F", "AB", "ABS", "DT"}
    for rec in records:
        roll = rec.get("roll_no", "")
        backs = back_map_for_sem.get(roll, {})
        if backs:
            matched += 1
        for s in rec.get("subjects", []):
            code = s["code"].replace(" ", "").upper()
            grade = (s.get("grade") or "").strip().upper()
            for bcode in backs:
                if bcode.replace(" ", "").upper() == code:
                    if grade in skip_grades:
                        # back marker still recorded for analytics, but no *.
                        s["back"] = True
                        s["back_pending"] = True
                        break
                    s["back"] = True
                    if not s["name"].rstrip().endswith("*"):
                        s["name"] = s["name"].rstrip() + " *"
                    n += 1
                    break
    if back_map_for_sem and matched == 0 and record_rolls:
        sample_excel = next(iter(back_map_for_sem.keys()))
        sample_pdf = next(iter(record_rolls))
        import logging as _log
        _log.getLogger("portal").warning(
            "apply_back_markers: 0 of %d Excel rolls matched any PDF roll. "
            "Excel sample=%s vs PDF sample=%s — Excel and PDF may belong to "
            "different batches.",
            len(back_map_for_sem), sample_excel, sample_pdf,
        )
    return n, matched


def restore_external_from_total(records: List[dict]) -> int:
    """B.Tech TC fix: when a subject's external column shows a dash but the
    total marks exceed the sessional marks, the external marks were
    suppressed/erased in the source data even though the student actually
    earned them. Recompute external = total - sessional.

    Marks are stored as ``"a/b"`` strings (a=obtained, b=max). We only
    restore when:
      - external is missing / dash-like ("", "-", "—" or "-/N" form), AND
      - sessional and total are valid ``"a/b"`` strings, AND
      - total_obtained > sessional_obtained  (the user's trigger condition).

    Non-credit subjects ($) are skipped — their external is intentionally
    a dash per institute policy.
    Returns the number of subjects whose external was restored.
    """
    fixed = 0
    pat = re.compile(r"^\s*(-?\d+)\s*/\s*(\d+)\s*$")
    for rec in records:
        for s in rec.get("subjects", []):
            if s.get("non_credit"):
                continue
            ext_raw = (s.get("external") or "").strip()
            # Treat plain "-", "—", "" or "-/N" / "/N" forms as missing.
            ext_a_present = bool(pat.match(ext_raw)) and not ext_raw.lstrip().startswith("-/")
            if ext_a_present:
                continue
            ses_m = pat.match((s.get("sessional") or "").strip())
            tot_m = pat.match((s.get("total") or "").strip())
            if not ses_m or not tot_m:
                continue
            try:
                ses_a = int(ses_m.group(1))
                ses_b = int(ses_m.group(2))
                tot_a = int(tot_m.group(1))
                tot_b = int(tot_m.group(2))
            except ValueError:
                continue
            if tot_a <= ses_a or tot_b <= ses_b:
                continue
            ext_a = tot_a - ses_a
            ext_b = tot_b - ses_b
            s["external"] = f"{ext_a}/{ext_b}"
            fixed += 1
    return fixed


def apply_non_credit_markers(records: List[dict], branch: str = "") -> int:
    """Append ' $' to subject names whose credits are 0 / blank — these are
    non-credit subjects (e.g. General Proficiency). Idempotent.

    By default, non-credit subjects don't carry an external (theory) exam,
    so external is forced to '-'. EXCEPTION: M.Tech Biotechnology's
    "English for Research Writing ($)" actually has an external exam — for
    this specific subject we keep whatever external marks the Excel
    supplied.
    """
    is_biotech = "biotech" in (branch or "").strip().lower()
    n = 0
    for rec in records:
        for s in rec.get("subjects", []):
            cred = (s.get("credits") or "").strip()
            try:
                is_zero = int(cred) == 0
            except Exception:
                is_zero = cred in ("", "-", "—", "0")
            if is_zero:
                s["non_credit"] = True
                if "$" not in s["name"]:
                    s["name"] = s["name"].rstrip(" *") + " $"
                    if s.get("back") and not s.get("back_pending"):
                        s["name"] += " *"
                    n += 1
                # Non-credit subjects don't carry an external (theory) exam —
                # only sessional / total / grade are awarded. The Excel may
                # still hold an "external" placeholder; replace it with a
                # dash so the printed TC reflects institute policy.
                # Biotechnology exception: "English for Research Writing"
                # does carry an external exam — keep original Excel value.
                name_norm = (s.get("name") or "").lower()
                keep_external = (
                    is_biotech
                    and "english for research writing" in name_norm
                )
                if not keep_external:
                    s["external"] = "-"
    return n


def _derive_grade_from_gp(gp_raw) -> str:
    """Map a numeric grade-point value to the equivalent letter grade.

    Uses GBPIET's standard 10-point credit-base scale (B.Tech / MCA):
    10 → O, 9 → A+, 8 → A, 7 → B+, 6 → B, 5 → C, 4 → D, <4 → F.
    Empty / non-numeric input returns ''.
    """
    if gp_raw in (None, ""):
        return ""
    try:
        gp = float(str(gp_raw).strip())
    except (TypeError, ValueError):
        return ""
    if gp >= 10: return "O"
    if gp >= 9:  return "A+"
    if gp >= 8:  return "A"
    if gp >= 7:  return "B+"
    if gp >= 6:  return "B"
    if gp >= 5:  return "C"
    if gp >= 4:  return "D"
    return "F"


def _derive_mtech_grade_from_gp(gp_raw) -> str:
    """Map a numeric grade-point value to the M.Tech ordinance letter grade.

    GBPIET M.Tech ordinance scale (see Grade Reference table on TC/GS):
        10 → O, 9 → A+, 8 → A, 7 → B, 6 → C, 5 → P, <5 → F.
    Note: M.Tech has NO B+ / D / E grades — those B.Tech buckets collapse.
    Empty / non-numeric input returns ''.
    """
    if gp_raw in (None, ""):
        return ""
    try:
        gp = float(str(gp_raw).strip())
    except (TypeError, ValueError):
        return ""
    if gp >= 10: return "O"
    if gp >= 9:  return "A+"
    if gp >= 8:  return "A"
    if gp >= 7:  return "B"
    if gp >= 6:  return "C"
    if gp >= 5:  return "P"
    return "F"


def fix_mtech_non_credit_columns(records: List[dict], branch: str) -> int:
    """For M.Tech branches OTHER than Biotechnology the Marksheets sheet
    lays out the non-credit "Technical Writing and Presentation Skill ($)"
    row with the sessional / total marks in the column my parser maps to
    "grade", and the letter grade missing entirely.

    Apply a corrective shift only for those branches:
      • external      → '-'
      • sessional, total → value found in current grade column (X/Y format)
      • grade         → derived from grade_points using the M.Tech
                        ordinance scale (10=O, 9=A+, 8=A, 7=B, 6=C, 5=P)

    Biotechnology keeps the existing layout untouched.
    """
    branch_norm = (branch or "").strip().lower()
    if "biotech" in branch_norm:
        return 0
    moved = 0
    for rec in records:
        for s in rec.get("subjects", []):
            if not s.get("non_credit"):
                continue
            grade_val = (s.get("grade") or "").strip()
            # Only shift when the grade column actually carries an "N/N"
            # marks-like value; if it already holds a real grade letter the
            # row was parsed correctly and we leave it alone.
            if "/" in grade_val:
                s["external"] = "-"
                s["sessional"] = grade_val
                s["total"] = grade_val
                s["grade"] = _derive_mtech_grade_from_gp(s.get("grade_points", ""))
                moved += 1
    return moved


# ---------------------------------------------------------------------------
# Generate PDF — TC*  (Tabulation Chart with asterisks)
# ---------------------------------------------------------------------------

_INSTITUTE_LINES = [
    "GOVIND BALLABH PANT INSTITUTE OF ENGINEERING AND TECHNOLOGY",
    "PAURI GARHWAL, UTTARAKHAND",
    "(AN AUTONOMOUS INSTITUTE OF GOVT. OF UTTARAKHAND)",
    "(AFFILIATED TO VEER MADHO SINGH BHANDARI UTTARAKHAND TECHNICAL UNIVERSITY)",
]

# Logo paths — extracted once from the original sample PDFs and bundled with the app.
_ASSETS = Path(__file__).parent / "assets"
INSTITUTE_LOGO = _ASSETS / "institute_logo.png"
UTU_LOGO = _ASSETS / "utu_logo.png"

# Optional decorative font for the institute name on the GS. If the TTF is
# present we register it with ReportLab so the font can be used directly via
# its PostScript-style name; otherwise we silently fall back to Helvetica-Bold.
_ALGERIAN_TTF = _ASSETS / "fonts" / "Algerian.ttf"
_GS_TITLE_FONT = "Helvetica-Bold"
try:
    if _ALGERIAN_TTF.exists():
        from reportlab.pdfbase import pdfmetrics as _pdfm
        from reportlab.pdfbase.ttfonts import TTFont as _TTFont
        _pdfm.registerFont(_TTFont("Algerian", str(_ALGERIAN_TTF)))
        _GS_TITLE_FONT = "Algerian"
except Exception:
    _GS_TITLE_FONT = "Helvetica-Bold"


from reportlab.platypus import Image as RLImage  # noqa: E402

# Barcode generation (Code 128) — used on every GS page
import barcode as _barcode_lib  # noqa: E402
from barcode.writer import ImageWriter  # noqa: E402


def _make_barcode_png(code_text: str) -> Optional[io.BytesIO]:
    """Generate a Code-128 barcode PNG for the given text. Returns BytesIO."""
    if not code_text:
        return None
    try:
        Code128 = _barcode_lib.get_barcode_class("code128")
        buf = io.BytesIO()
        # write_text=False keeps the image clean; we render text below ourselves
        Code128(code_text, writer=ImageWriter()).write(
            buf,
            options={"module_width": 0.30, "module_height": 8.0, "write_text": False, "quiet_zone": 1.0},
        )
        buf.seek(0)
        return buf
    except Exception:
        return None


def _logo_header(text_lines, total_width_mm: float, logo_size_mm: float = 22):
    """Build a 3-column row: [institute logo] [stacked text lines] [UTU logo]."""
    text_cell = []
    for txt, style in text_lines:
        text_cell.append(Paragraph(txt, style))

    inst = (
        RLImage(str(INSTITUTE_LOGO), width=logo_size_mm * mm, height=logo_size_mm * mm)
        if INSTITUTE_LOGO.exists()
        else ""
    )
    utu = (
        RLImage(str(UTU_LOGO), width=logo_size_mm * mm, height=logo_size_mm * mm)
        if UTU_LOGO.exists()
        else ""
    )
    text_w = total_width_mm - 2 * (logo_size_mm + 2)
    t = Table(
        [[inst, text_cell, utu]],
        colWidths=[(logo_size_mm + 2) * mm, text_w * mm, (logo_size_mm + 2) * mm],
    )
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    return t


def _styles():
    s = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", parent=s["Heading1"], fontName="Helvetica-Bold",
                                 fontSize=11, alignment=1, spaceAfter=2, leading=13),
        "sub": ParagraphStyle("sub", parent=s["Normal"], fontName="Helvetica",
                               fontSize=9, alignment=1, spaceAfter=1, leading=11),
        "small": ParagraphStyle("small", parent=s["Normal"], fontName="Helvetica",
                                 fontSize=8, alignment=1, spaceAfter=1, leading=10),
        "section": ParagraphStyle("section", parent=s["Normal"], fontName="Helvetica-Bold",
                                   fontSize=11, alignment=1, spaceBefore=4, spaceAfter=4,
                                   leading=13),
        "label": ParagraphStyle("label", parent=s["Normal"], fontName="Helvetica",
                                 fontSize=9, alignment=0, leading=11),
        "right": ParagraphStyle("right", parent=s["Normal"], fontName="Helvetica",
                                 fontSize=9, alignment=2),
        "back_subject": ParagraphStyle("back_subject", parent=s["Normal"],
                                        fontName="Helvetica-Bold", fontSize=8,
                                        textColor=colors.HexColor("#92400e"), alignment=0,
                                        leading=10),
        "subject": ParagraphStyle("subject", parent=s["Normal"], fontName="Helvetica",
                                    fontSize=8, alignment=0, leading=10),
    }


def _draw_tc_header(canv, doc, program: str, branch: str, sem: str, session: str):
    """Per-page header for TC: institute logos + name strip + course/branch/semester.

    Drawn directly on the canvas so it appears on every A3 page. Sized for A3
    (297×420 mm) — logos 30 mm, institute name 15 pt bold.
    """
    canv.saveState()
    page_w, _page_h = doc.pagesize
    margin = 12 * mm
    logo_size = 30 * mm
    top_y = _page_h - 6 * mm  # top edge of logo

    # Logos (left = institute, right = UTU) — silently skip if files missing.
    if INSTITUTE_LOGO.exists():
        try:
            canv.drawImage(
                str(INSTITUTE_LOGO), margin, top_y - logo_size,
                width=logo_size, height=logo_size,
                mask="auto", preserveAspectRatio=True,
            )
        except Exception:
            pass
    if UTU_LOGO.exists():
        try:
            canv.drawImage(
                str(UTU_LOGO), page_w - margin - logo_size, top_y - logo_size,
                width=logo_size, height=logo_size,
                mask="auto", preserveAspectRatio=True,
            )
        except Exception:
            pass

    # Centered institute text block (4 lines beside logos)
    cx = page_w / 2
    canv.setFillColor(colors.black)
    canv.setFont("Helvetica-Bold", 15)
    canv.drawCentredString(cx, top_y - 6 * mm, _INSTITUTE_LINES[0])
    canv.setFont("Helvetica", 11.5)
    canv.drawCentredString(cx, top_y - 11.5 * mm, _INSTITUTE_LINES[1])
    canv.setFont("Helvetica", 10)
    canv.drawCentredString(cx, top_y - 16.5 * mm, _INSTITUTE_LINES[2])
    canv.drawCentredString(cx, top_y - 21 * mm, _INSTITUTE_LINES[3])

    # Course / branch / semester block — sits just below logos, full width.
    course_y = top_y - logo_size - 5 * mm
    canv.setFont("Helvetica-Bold", 14)
    canv.drawCentredString(
        cx, course_y, f"Tabulation Chart for {program_full(program)}"
    )
    branch_disp = branch_full(branch)
    # For MCA, branch == programme name (single branch) — skip the duplicate
    # branch line so the course name doesn't appear twice on the TC.
    show_branch = bool(branch_disp) and program_short(program) != "MCA"
    if show_branch:
        canv.setFont("Helvetica-Bold", 12)
        canv.drawCentredString(cx, course_y - 5.5 * mm, branch_disp)
        sess_y_offset = 10.5 * mm
    else:
        sess_y_offset = 6 * mm
    canv.setFont("Helvetica", 11)
    sess_line = f"{sem} Semester"
    if session:
        sess_line += f"   |   {session}"
    canv.drawCentredString(cx, course_y - sess_y_offset, sess_line)

    # Thin divider beneath header
    canv.setLineWidth(0.5)
    canv.setStrokeColor(colors.HexColor("#9ca3af"))
    canv.line(margin, course_y - 13.5 * mm, page_w - margin, course_y - 13.5 * mm)
    canv.restoreState()


def _draw_tc_footer(canv, doc, program: str, branch: str, sem: str, session: str):
    """Per-page footer for TC: 5-cell signature row + page number.
    (Legend is rendered in-flow, immediately after the last student record.)"""
    canv.saveState()
    page_w, _ = doc.pagesize
    margin = 12 * mm
    # ---- Signature row ----
    y = 14 * mm
    cell_w = (page_w - 2 * margin) / 5.0
    labels = [
        "Prepared by", "Checked by",
        "Examination Controller", "Director",
        "Examination Controller (VMSB UTU)",
    ]
    canv.setLineWidth(0.4)
    canv.setStrokeColor(colors.HexColor("#9ca3af"))
    canv.line(margin, y + 2, page_w - margin, y + 2)
    canv.setFont("Helvetica", 9)
    canv.setFillColor(colors.HexColor("#1c1917"))
    for i, lbl in enumerate(labels):
        x = margin + i * cell_w + cell_w / 2
        canv.drawCentredString(x, y - 6, lbl)
    canv.setFont("Helvetica", 7)
    canv.setFillColor(colors.HexColor("#57534e"))
    canv.drawRightString(page_w - margin, 6 * mm, f"Page {doc.page}")
    canv.restoreState()


def _draw_tc_page(canv, doc, program: str, branch: str, sem: str, session: str):
    """Combined per-page chrome: header (top) + footer (bottom)."""
    _draw_tc_header(canv, doc, program, branch, sem, session)
    _draw_tc_footer(canv, doc, program, branch, sem, session)


def generate_tc_pdf(records: List[dict], program: str = "", branch: str = "",
                     semester_roman: str = "", exam_session: str = "") -> bytes:
    """Tabulation Chart — A3 portrait, 4 students per page max, NO content
    spanning two pages (KeepTogether). Per-page signed footer. No grade
    reference. Programme / branch / semester / session are taken from the
    sheet-level metadata stored on each record (with the function arguments
    only as fallback)."""
    # Resolve metadata from the records themselves where possible.
    if records:
        meta = records[0]
        program = meta.get("program", "") or program
        branch = meta.get("branch", "") or branch
        semester_roman = meta.get("semester", "") or semester_roman
        exam_session = meta.get("exam_session", "") or exam_session

    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, KeepTogether, Flowable,
    )

    buf = io.BytesIO()
    page = A3
    margin = 12 * mm
    footer_h = 28 * mm  # signature band
    header_h = 60 * mm  # institute strip + course/branch/semester (drawn on canvas)

    # ---- Per-page legend, drawn dynamically just below the last student block ----
    last_block_bottom: dict[int, float] = {}

    class _RecordY(Flowable):
        """Zero-height marker; records the absolute y position where it draws.

        We append one of these after every student block; the LAST one drawn on a
        given page reveals where the bottom of that page's content sits, so the
        legend can be placed immediately beneath it (rather than floating in
        whitespace at the bottom of the page).
        """
        def wrap(self, _w, _h):
            return (1, 0)

        def draw(self):
            _x, y = self.canv.absolutePosition(0, 0)
            last_block_bottom[self.canv.getPageNumber()] = y

    def _on_page_end(c, _d):
        page_num = c.getPageNumber()
        y = last_block_bottom.get(page_num)
        if y is None:
            return
        c.saveState()
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColor(colors.HexColor("#57534e"))
        c.drawString(
            margin, y - 4.5 * mm,
            "*  subject cleared after back paper          $  non-credit subject",
        )
        c.restoreState()

    doc = BaseDocTemplate(
        buf, pagesize=page,
        leftMargin=margin, rightMargin=margin,
        topMargin=header_h, bottomMargin=footer_h,
    )
    frame = Frame(
        margin, footer_h, page[0] - 2 * margin, page[1] - header_h - footer_h,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        showBoundary=0,
    )
    doc.addPageTemplates([PageTemplate(
        id="tc", frames=[frame],
        onPage=lambda c, d: _draw_tc_page(c, d, program, branch, semester_roman, exam_session),
        onPageEnd=_on_page_end,
    )])

    st = _styles()
    page_width_mm = page[0] / mm - 24
    col_widths = [26 * mm, 95 * mm, 16 * mm, 28 * mm, 28 * mm, 28 * mm, 22 * mm, 26 * mm]

    story: list = []
    students_per_page = 4

    def build_student_block(rec):
        info = [[
            Paragraph(f"<b>Name:</b> {rec.get('name','')}", st["label"]),
            Paragraph(f"<b>Father's Name:</b> {rec.get('father_name','')}", st["label"]),
            Paragraph(f"<b>University Roll No.:</b> {rec.get('roll_no','')}", st["label"]),
            Paragraph(f"<b>University Enrol. No.:</b> {rec.get('enroll_no','')}", st["label"]),
        ]]
        t_info = Table(info, colWidths=[70 * mm, 70 * mm, 65 * mm, 64 * mm])
        t_info.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        rows = [[
            "Subject\nCode", "Subject Name", "Credits", "External\nMarks",
            "Sessional\nMarks", "Total\nMarks", "Grade", "Grade\nPoints",
        ]]
        for s in rec.get("subjects", []):
            # Only highlight subjects that were highlighted in SEM_ sheet AND
            # the student actually cleared the back paper (grade not F/Ab/Dt).
            cleared_back = s.get("back") and not s.get("back_pending")
            name_para = Paragraph(s["name"], st["back_subject"] if cleared_back else st["subject"])
            rows.append([
                s["code"], name_para, s.get("credits", ""), s.get("external", ""),
                s.get("sessional", ""), s.get("total", ""), s.get("grade", ""),
                s.get("grade_points", ""),
            ])
        totals = _compute_totals(rec.get("subjects", []))
        rows.append([
            Paragraph("<b>Total Credits / Marks / Total Grade Points</b>", st["label"]),
            "", str(totals["credits"]), totals["external"], totals["sessional"],
            totals["total"], "", f"{totals['grade_points']:.1f}",
        ])
        rows.append([
            Paragraph(f"<b>Result:</b> {rec.get('result','—')}", st["label"]), "",
            Paragraph(f"<b>Remark:</b> {rec.get('remark','—') or '—'}", st["label"]), "",
            Paragraph(f"<b>SGPA:</b> {rec.get('sgpa','—')}", st["label"]), "",
            Paragraph(f"<b>CGPA:</b> {rec.get('cgpa','—')}", st["label"]), "",
        ])
        rows.append([
            Paragraph(f"<b>Earned Credits:</b> {rec.get('earned_credits','—')}", st["label"]),
            "", "", "",
            Paragraph(
                f"<b>Cumulative Earned Credits:</b> "
                f"{rec.get('cuml_earned_credits') or rec.get('earned_credits','—')}",
                st["label"],
            ), "", "", "",
        ])
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
            ("GRID", (0, 0), (-1, -4), 0.25, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e1b4b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (2, 1), (-1, -4), "CENTER"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("SPAN", (0, -3), (1, -3)),
            ("BACKGROUND", (0, -3), (-1, -3), colors.HexColor("#f5f5f4")),
            ("FONTNAME", (0, -3), (-1, -3), "Helvetica-Bold"),
            ("ALIGN", (2, -3), (-1, -3), "CENTER"),
            ("SPAN", (0, -2), (1, -2)),
            ("SPAN", (2, -2), (3, -2)),
            ("SPAN", (4, -2), (5, -2)),
            ("SPAN", (6, -2), (7, -2)),
            ("FONTNAME", (0, -2), (-1, -2), "Helvetica-Bold"),
            ("BACKGROUND", (0, -2), (-1, -2), colors.HexColor("#fef3c7")),
            ("SPAN", (0, -1), (3, -1)),
            ("SPAN", (4, -1), (7, -1)),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fafaf9")),
        ]))
        return KeepTogether([t_info, t, Spacer(1, 2.5 * mm)])

    # Auto-fit: ReportLab places as many KeepTogether blocks per page as fit.
    # After every block we drop a zero-height _RecordY marker so the
    # onPageEnd hook can draw the legend just below the page's last block.
    for rec in records:
        story.append(build_student_block(rec))
        story.append(_RecordY())

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Generate PDF — GS*  (Grade Sheet with asterisks + barcode)
# ---------------------------------------------------------------------------


def _draw_gs_decoration(canv, doc, watermark_path: str = ""):
    """Per-page decorations for GS: faint institute-logo watermark drawn
    BENEATH the flowables. Kept as a softer base; the more visible watermark
    is drawn in `_draw_gs_watermark_on_top` via onPageEnd."""
    canv.saveState()
    page_w, page_h = doc.pagesize
    if watermark_path and Path(watermark_path).exists():
        try:
            canv.setFillAlpha(0.05)
            wmark_w = 130 * mm
            wmark_h = 130 * mm
            canv.drawImage(
                watermark_path,
                (page_w - wmark_w) / 2, (page_h - wmark_h) / 2,
                width=wmark_w, height=wmark_h, mask="auto",
                preserveAspectRatio=True,
            )
            canv.setFillAlpha(1)
        except Exception:
            pass
    canv.restoreState()


def _draw_gs_signature_footer(canv, doc):
    """Per-page signature row at the bottom of every GS page (Prepared by,
    Checked by, Examination Controller, Director). Drawn directly on the
    canvas so it appears even when the per-student story doesn't reach the
    bottom of the frame."""
    canv.saveState()
    page_w, _page_h = doc.pagesize
    margin = 14 * mm
    y = 14 * mm
    cell_w = (page_w - 2 * margin) / 4.0
    labels = [
        "Prepared by", "Checked by", "Examination Controller", "Director",
    ]
    canv.setLineWidth(0.5)
    canv.setStrokeColor(colors.HexColor("#9ca3af"))
    canv.line(margin, y + 4, page_w - margin, y + 4)
    canv.setFont("Helvetica-Bold", 9.5)
    canv.setFillColor(colors.HexColor("#1c1917"))
    for i, lbl in enumerate(labels):
        x = margin + i * cell_w + cell_w / 2
        canv.drawCentredString(x, y - 4, lbl)
    canv.restoreState()


def _draw_gs_header(canv, doc):
    """Draws the institute strip + side logos directly on the GS page canvas.

    The institute name auto-shrinks to fit a single line within the available
    width between the two logos. Subsequent affiliation lines render at a
    fixed smaller size."""
    from reportlab.pdfbase.pdfmetrics import stringWidth
    canv.saveState()
    page_w, page_h = doc.pagesize
    margin = 14 * mm
    logo_size = 22 * mm
    top_y = page_h - 8 * mm  # top edge of logo (extra padding from page top)

    # Logos
    if INSTITUTE_LOGO.exists():
        try:
            canv.drawImage(
                str(INSTITUTE_LOGO), margin, top_y - logo_size,
                width=logo_size, height=logo_size,
                mask="auto", preserveAspectRatio=True,
            )
        except Exception:
            pass
    if UTU_LOGO.exists():
        try:
            canv.drawImage(
                str(UTU_LOGO), page_w - margin - logo_size, top_y - logo_size,
                width=logo_size, height=logo_size,
                mask="auto", preserveAspectRatio=True,
            )
        except Exception:
            pass

    cx = page_w / 2
    canv.setFillColor(colors.black)

    # Available text width between the logos (with 4mm padding either side).
    text_w = page_w - 2 * margin - 2 * (logo_size + 4 * mm)
    name_text = _INSTITUTE_LINES[0]
    # Auto-shrink the institute name to the largest size that fits in one line
    # using the configured GS title font (Algerian if available, else
    # Helvetica-Bold). Algerian glyphs are wider so we widen the search range.
    name_size = 16.0
    while name_size > 7 and stringWidth(name_text, _GS_TITLE_FONT, name_size) > text_w:
        name_size -= 0.25

    canv.setFont(_GS_TITLE_FONT, name_size)
    canv.drawCentredString(cx, top_y - 5 * mm, name_text)
    # Affiliation lines also use the GS title font for visual consistency.
    canv.setFont(_GS_TITLE_FONT, 9.5)
    canv.drawCentredString(cx, top_y - 10 * mm, _INSTITUTE_LINES[1])
    canv.setFont(_GS_TITLE_FONT, 8)
    canv.drawCentredString(cx, top_y - 14 * mm, _INSTITUTE_LINES[2])
    canv.drawCentredString(cx, top_y - 17.5 * mm, _INSTITUTE_LINES[3])
    canv.restoreState()


def _draw_gs_watermark_on_top(canv, doc, watermark_path: str = ""):
    """Draws the institute-logo watermark ON TOP of the page content with a
    very low alpha so it's perceivable yet keeps the table text legible.
    Called via onPageEnd after the frame has rendered."""
    if not watermark_path or not Path(watermark_path).exists():
        return
    canv.saveState()
    page_w, page_h = doc.pagesize
    try:
        canv.setFillAlpha(0.10)
        wmark_w = 140 * mm
        wmark_h = 140 * mm
        canv.drawImage(
            watermark_path,
            (page_w - wmark_w) / 2, (page_h - wmark_h) / 2,
            width=wmark_w, height=wmark_h, mask="auto",
            preserveAspectRatio=True,
        )
        canv.setFillAlpha(1)
    except Exception:
        pass
    canv.restoreState()


def generate_gs_pdf(records: List[dict], program: str = "", branch: str = "",
                     semester_roman: str = "", exam_session: str = "",
                     batch: str = "",
                     all_sem_summary: Optional[Dict[str, Dict[str, Dict[str, str]]]] = None) -> bytes:
    """Polished A4-portrait Grade Sheet.

    `all_sem_summary` (optional): {roll_no: {SEM: {sgpa, cgpa, earned_credits,
    result}}} — used to render the per-semester SGPA/CGPA history table at
    the bottom of every page.
    """
    if records:
        meta = records[0]
        program = meta.get("program", "") or program
        branch = meta.get("branch", "") or branch
        semester_roman = meta.get("semester", "") or semester_roman
        exam_session = meta.get("exam_session", "") or exam_session

    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, KeepTogether,
    )

    buf = io.BytesIO()
    page = A4
    margin = 14 * mm
    bottom_h = 26 * mm  # leave room for the canvas-drawn signature footer
    top_h = 38 * mm     # leave generous room for the canvas-drawn institute header
    doc = BaseDocTemplate(
        buf, pagesize=page,
        leftMargin=margin, rightMargin=margin,
        topMargin=top_h, bottomMargin=bottom_h,
    )
    frame = Frame(
        margin, bottom_h, page[0] - 2 * margin, page[1] - top_h - bottom_h,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        showBoundary=0,
    )

    def _gs_on_page(c, d):
        _draw_gs_decoration(c, d, str(INSTITUTE_LOGO))
        _draw_gs_header(c, d)
        _draw_gs_signature_footer(c, d)

    def _gs_on_page_end(c, d):
        _draw_gs_watermark_on_top(c, d, str(INSTITUTE_LOGO))

    doc.addPageTemplates([PageTemplate(
        id="gs", frames=[frame],
        onPage=_gs_on_page,
        onPageEnd=_gs_on_page_end,
    )])

    st = _styles()
    page_width_mm = page[0] / mm - 2 * (margin / mm)
    story: list = []

    # Map of {roll_no: gs_hash} so the caller can persist verification codes.
    hash_map: Dict[str, str] = {}

    # KeepInFrame wraps every student's GS into a single-page block; if the
    # combined content overflows, ReportLab shrinks fonts/spacing
    # proportionally so the entire GS still fits on ONE A4 page.
    from reportlab.platypus import KeepInFrame
    frame_w = page[0] - 2 * margin
    frame_h = page[1] - top_h - bottom_h

    for idx, rec in enumerate(records):
        _start = len(story)  # remember where this student's flowables begin
        # ---- Verification hash (12 hex chars) — content-addressed.
        # Seeds from every visible GS field (subjects, grades, marks, totals,
        # asterisks via "back" flag, etc.) so that ANY change to the GS — a
        # newly-applied asterisk, an updated grade or remark — produces a
        # different hash on the next upload, while an unchanged GS keeps
        # exactly the same code.
        import hashlib, json
        subj_seed = json.dumps(
            [
                {
                    "code": s.get("code", ""),
                    "name": s.get("name", ""),
                    "credits": s.get("credits", ""),
                    "external": s.get("external", ""),
                    "sessional": s.get("sessional", ""),
                    "total": s.get("total", ""),
                    "grade": s.get("grade", ""),
                    "grade_points": s.get("grade_points", ""),
                    "back": bool(s.get("back")),
                    "back_pending": bool(s.get("back_pending")),
                }
                for s in rec.get("subjects", [])
            ],
            sort_keys=True,
            ensure_ascii=False,
        )
        seed = "|".join([
            rec.get("roll_no", ""), rec.get("enroll_no", ""),
            rec.get("name", ""), rec.get("father_name", ""),
            program, branch, batch,
            semester_roman, exam_session,
            str(rec.get("sgpa", "")), str(rec.get("cgpa", "")),
            str(rec.get("earned_credits", "")),
            str(rec.get("cuml_earned_credits", "")),
            rec.get("result", ""), rec.get("remark", ""),
            subj_seed,
        ])
        verify_hash = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12].upper()
        rec["gs_hash"] = verify_hash
        if rec.get("roll_no"):
            hash_map[rec["roll_no"]] = verify_hash
        bc_payload = f"{rec.get('roll_no','')}-{verify_hash}"

        # ---- Top-right barcode (replaces SL.NO.) ----
        bc_top = _make_barcode_png(bc_payload)
        if bc_top is not None:
            try:
                top_bc = RLImage(bc_top, width=52 * mm, height=11 * mm)
                hash_p = Paragraph(
                    f"<font size='8' face='Helvetica-Bold' color='#1c1917'>"
                    f"{verify_hash}</font>",
                    st["small"],
                )
                top_table = Table(
                    [[ "", top_bc ],
                     [ "", hash_p ]],
                    colWidths=[(page_width_mm - 56) * mm, 56 * mm],
                )
                top_table.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]))
                story.append(top_table)
            except Exception:
                pass

        # Institute strip is rendered by the canvas via _draw_gs_header on
        # every page, so the story starts directly with the GRADE SHEET title
        # band below.

        # Common content width — both header bands AND the tables below use
        # this exact width so all sections align flush left/right.
        content_w = 174 * mm

        # ---- "GRADE SHEET" header band — its own visual section ----
        title_band = Table([[Paragraph(
            "<para alignment='center' leading='16' spaceBefore='0' spaceAfter='0'>"
            "<font size='14' face='Helvetica-Bold' color='white'>GRADE SHEET</font>"
            "</para>",
            st["sub"],
        )]], colWidths=[content_w])
        title_band.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1e1b4b")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(title_band)
        story.append(Spacer(1, 1.5 * mm))

        # ---- Course / branch / semester+session — its own section ----
        # MCA has no per-branch breakdown, so suppress the redundant branch
        # line whenever the programme is MCA (the branch field on those records
        # mirrors the programme label).
        course_full = program_full(program)
        branch_disp = branch_full(branch)
        show_branch = bool(branch_disp) and program_short(program) != "MCA"
        course_html = (
            f"<font size='13' face='Times-Bold' color='#1c1917'>{course_full}</font><br/>"
        )
        if show_branch:
            course_html += (
                f"<font size='11.5' face='Times-Roman' color='#1c1917'>{branch_disp}</font><br/>"
            )
        course_html += (
            f"<font size='10.5' face='Times-Italic' color='#57534e'>"
            f"{semester_roman} Semester &nbsp;&middot;&nbsp; {exam_session}</font>"
        )
        course_band = Table([[Paragraph(
            f"<para alignment='center' leading='17'>{course_html}</para>",
            st["sub"],
        )]], colWidths=[content_w])
        course_band.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f5f5f4")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(course_band)
        story.append(Spacer(1, 3 * mm))

        # ---- Student details card ----
        info = Table([
            [Paragraph("<font size='7.5' color='#57534e'><b>NAME</b></font>", st["small"]),
             Paragraph(f"<font size='10.5'><b>{rec.get('name','')}</b></font>", st["label"]),
             Paragraph("<font size='7.5' color='#57534e'><b>UNIVERSITY ROLL NO.</b></font>", st["small"]),
             Paragraph(f"<font size='10.5' face='Helvetica-Bold'>{rec.get('roll_no','')}</font>",
                       st["label"])],
            [Paragraph("<font size='7.5' color='#57534e'><b>FATHER'S NAME</b></font>", st["small"]),
             Paragraph(f"<font size='10'><b>{rec.get('father_name','')}</b></font>", st["label"]),
             Paragraph("<font size='7.5' color='#57534e'><b>UNIVERSITY ENROLL. NO.</b></font>", st["small"]),
             Paragraph(f"<font size='10' face='Helvetica-Bold'>{rec.get('enroll_no','')}</font>", st["label"])],
        ], colWidths=[34 * mm, 60 * mm, 40 * mm, 40 * mm])
        info.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f5f5f4")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f5f5f4")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(info)
        story.append(Spacer(1, 3 * mm))

        # ---- Subjects table ----
        # GS-specific subject paragraph styles — sized to match the surrounding
        # table cells (codes/grades render at 9 pt via the table FONTSIZE rule),
        # so the wrapped subject names visually align with the rest of the row.
        gs_subject_style = ParagraphStyle(
            "gs_subject", parent=st["label"], fontName="Helvetica",
            fontSize=9, alignment=0, leading=11,
        )
        gs_back_subject_style = ParagraphStyle(
            "gs_back_subject", parent=st["label"], fontName="Helvetica-Bold",
            fontSize=9, alignment=0, leading=11,
            textColor=colors.HexColor("#92400e"),
        )
        rows = [["Subject Code", "Subject Name", "Credits", "Grade", "Grade Points"]]
        total_credits = 0
        total_gp = 0.0
        for s in rec.get("subjects", []):
            cleared_back = s.get("back") and not s.get("back_pending")
            name_para = Paragraph(
                s["name"],
                gs_back_subject_style if cleared_back else gs_subject_style,
            )
            rows.append([s["code"], name_para, s.get("credits", ""),
                         s.get("grade", ""), s.get("grade_points", "")])
            try:
                total_credits += int(s.get("credits") or 0)
            except Exception:
                pass
            try:
                total_gp += float(s.get("grade_points") or 0) * int(s.get("credits") or 0)
            except Exception:
                pass
        rows.append([
            Paragraph("<b>Total Credits / Total Grade Points</b>", st["label"]),
            "", str(total_credits), "—", f"{total_gp:.1f}",
        ])
        col_widths = [27 * mm, 78 * mm, 22 * mm, 22 * mm, 25 * mm]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e1b4b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, 0), 9.5),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2),
              [colors.white, colors.HexColor("#fafaf9")]),
            ("SPAN", (0, -1), (1, -1)),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef3c7")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 3 * mm))

        # ---- Semester-wise Result — full grid, future semesters blank ----
        # MCA and M.Tech are 4-semester programmes; B.Tech is 8.
        roll = rec.get("roll_no", "")
        per_sem = (all_sem_summary or {}).get(roll, {})
        if is_short_program(program):
            sem_order = ["I", "II", "III", "IV"]
        else:
            sem_order = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII"]
        try:
            current_idx = sem_order.index(semester_roman)
        except ValueError:
            current_idx = len(sem_order) - 1
        hist_header = ["Semester"] + sem_order
        sgpa_row = ["SGPA"]
        cgpa_row = ["CGPA"]
        ec_row = ["Earned Cr."]
        res_row = ["Result"]
        for i, s in enumerate(sem_order):
            if i > current_idx:
                # Future semester — leave blank/dash on the GS
                sgpa_row.append("—")
                cgpa_row.append("—")
                ec_row.append("—")
                res_row.append("—")
                continue
            cell = per_sem.get(s, {})
            if not cell and s == semester_roman:
                cell = {
                    "sgpa": rec.get("sgpa", ""),
                    "cgpa": rec.get("cgpa", ""),
                    "earned_credits": rec.get("earned_credits", ""),
                    "result": rec.get("result", ""),
                }
            sgpa_row.append(cell.get("sgpa") or "—")
            cgpa_row.append(cell.get("cgpa") or "—")
            ec_row.append(cell.get("earned_credits") or "—")
            res_row.append((cell.get("result") or "—")[:6])
        n_cols = len(sem_order)
        col_w = (page_width_mm - 24) / max(n_cols, 1)
        history = Table([hist_header, sgpa_row, cgpa_row, ec_row, res_row],
                          colWidths=[24 * mm] + [col_w * mm] * n_cols,
                          repeatRows=1)
        # Style: shade future-semester columns lighter so the eye knows they
        # belong to a later GS.
        future_col_style = []
        for i in range(current_idx + 2, len(sem_order) + 1):
            future_col_style.append(
                ("BACKGROUND", (i, 1), (i, -1), colors.HexColor("#fafaf9"))
            )
            future_col_style.append(
                ("TEXTCOLOR", (i, 1), (i, -1), colors.HexColor("#a8a29e"))
            )
        history.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e1b4b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f5f5f4")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            *future_col_style,
        ]))
        # Legend immediately above the semester-wise result table.
        story.append(Paragraph(
            "<font size='8' color='#57534e'>"
            "<b>*</b> &mdash; subject cleared after back paper "
            "&nbsp;&nbsp;&nbsp;&nbsp;"
            "<b>$</b> &mdash; non-credit subject"
            "</font>",
            st["label"],
        ))
        story.append(Spacer(1, 1.5 * mm))
        story.append(Paragraph(
            "<font size='8.5' color='#57534e'><b>SEMESTER-WISE RESULT</b></font>",
            st["label"],
        ))
        story.append(Spacer(1, 1 * mm))
        story.append(history)
        story.append(Spacer(1, 3 * mm))

        # ---- Summary card (SGPA / CGPA / Earned credits / Result) ----
        sgpa = rec.get("sgpa") or "—"
        cgpa = rec.get("cgpa") or "—"
        earned = rec.get("earned_credits") or str(total_credits)
        result_val = rec.get("result") or "—"
        remark_val = rec.get("remark") or ""
        # Custom wide-leading style for stat cells so the label and value
        # don't overlap (label 7.5pt → value 15pt requires ~9pt + 17pt of
        # vertical room). Using <para leading> on a flowable text guarantees
        # proper line spacing regardless of the parent style's leading.
        def _stat_cell(label, value):
            return Paragraph(
                "<para alignment='center' leading='14' spaceBefore='0' spaceAfter='0'>"
                f"<font size='7' color='#57534e'>{label}</font><br/>"
                f"<font size='10' face='Helvetica-Bold'>{value}</font>"
                "</para>",
                st["label"],
            )

        summary_top = Table([[
            _stat_cell("SGPA", sgpa),
            _stat_cell("CGPA", cgpa),
            _stat_cell("EARNED CREDITS", earned),
            _stat_cell("RESULT", result_val),
        ]], colWidths=[(page_width_mm / 4) * mm] * 4)
        summary_top.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafaf9")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(summary_top)
        if remark_val:
            story.append(Spacer(1, 1.5 * mm))
            story.append(Paragraph(
                f"<font size='8' color='#57534e'><b>REMARK:</b></font> "
                f"<font size='9'>{remark_val}</font>",
                st["label"],
            ))

        # Signature row is rendered as a per-page canvas footer
        # (`_draw_gs_signature_footer`), so we do NOT add it to the story.

        # ---- Single-page guarantee: wrap this student's flowables in a
        # KeepInFrame so they always fit one A4 page (auto-shrink on overflow).
        student_block = story[_start:]
        del story[_start:]
        story.append(KeepInFrame(
            frame_w, frame_h, student_block, mode="shrink",
            mergeSpace=1,
        ))

        if idx + 1 < len(records):
            story.append(PageBreak())

    doc.build(story)
    return buf.getvalue()


def _branch_short(branch: str) -> str:
    if not branch:
        return "XX"
    parts = re.findall(r"[A-Z]", branch.upper())
    return "".join(parts)[:4] or "XX"


def program_short(program: str) -> str:
    """Return canonical short label for a programme.

    Accepts long names ('Bachelor of Technology', 'Master of Computer
    Applications', 'Master of Technology') as well as the short literal forms
    used by the frontend ('B.Tech', 'M.Tech', 'MCA'). Falls back to 'PROG'.
    """
    if not program:
        return "PROG"
    norm = re.sub(r"[^A-Z]", "", program.upper())  # strip dots/spaces
    if "MASTEROFCOMPUTER" in norm or norm == "MCA":
        return "MCA"
    if "MASTEROFTECH" in norm or norm.startswith("MTECH"):
        return "M. TECH"
    if "BACHELOR" in norm or norm.startswith("BTECH"):
        return "B. TECH"
    if "MASTER" in norm:
        return "M. TECH"
    return "PROG"


def program_full(program: str) -> str:
    """Return the canonical FULL human-readable course name.

    Input may be either an abbreviation ('B.Tech', 'MCA', 'M.Tech') or
    something already verbose. Always emits one of the three official course
    titles used on every printed Grade Sheet / Tabulation Chart.
    """
    code = program_short(program)
    if code == "B. TECH":
        return "Bachelor of Technology"
    if code == "MCA":
        return "Master of Computer Application"
    if code == "M. TECH":
        return "Master of Technology"
    # Unknown / unrecognised — keep whatever the caller passed in so we don't
    # silently mask data, and fall back to a humanised version.
    return program or "Programme"


def is_short_program(program: str) -> bool:
    """True when the programme has only 4 semesters (MCA, M.Tech)."""
    return program_short(program) in ("MCA", "M. TECH")


# Common abbreviations seen in branch names that the institute prefers to
# always print in expanded form on official transcripts.
_BRANCH_ABBR_EXPANSIONS: List[Tuple[str, str]] = [
    ("AIML", "Artificial Intelligence and Machine Learning"),
    ("AI/ML", "Artificial Intelligence and Machine Learning"),
    ("AI ML", "Artificial Intelligence and Machine Learning"),
    ("AI&ML", "Artificial Intelligence and Machine Learning"),
    ("CSE", "Computer Science & Engineering"),
    ("ECE", "Electronics and Communication Engineering"),
    ("EEE", "Electrical and Electronics Engineering"),
    ("EE", "Electrical Engineering"),
    ("ME", "Mechanical Engineering"),
    ("CE", "Civil Engineering"),
    ("IT", "Information Technology"),
    ("BT", "Biotechnology"),
    ("CHE", "Chemical Engineering"),
    ("PROD", "Production Engineering"),
    ("THERMAL", "Thermal Engineering"),
    ("POWER SYSTEM", "Power System"),
    ("POWERSYSTEM", "Power System"),
    ("GEOTECHNOLOGY", "Geotechnology"),
    ("GEO TECH", "Geotechnology"),
]


def branch_full(branch: str) -> str:
    """Return the branch name with common short forms expanded.

    Operates on whole-word, case-insensitive boundaries. If a parenthesised
    short form is found (e.g. ``Computer Science & Engineering (AIML)``), only
    the parenthetical content is expanded so the existing prefix stays
    untouched.
    """
    if not branch:
        return ""
    out = branch
    # First expand parenthesised short forms: '(AIML)' → '(Artificial...)'
    def _expand_paren(m: "re.Match") -> str:
        inside = m.group(1).strip()
        for short, full in _BRANCH_ABBR_EXPANSIONS:
            if inside.upper() == short.upper():
                return f"({full})"
        return m.group(0)
    out = re.sub(r"\(([^()]+)\)", _expand_paren, out)
    # Then expand whole-word abbreviations elsewhere in the string.
    for short, full in _BRANCH_ABBR_EXPANSIONS:
        # Only expand if the short form appears as a whole word AND the full
        # form is not already present nearby (avoid replacing 'CSE' inside
        # 'Computer Science & Engineering').
        pattern = re.compile(rf"\b{re.escape(short)}\b", re.IGNORECASE)
        if pattern.search(out) and full.lower() not in out.lower():
            out = pattern.sub(full, out, count=1)
    return out


# ---------------------------------------------------------------------------
# Grade tables — reference tables embedded at the bottom of TC* / GS*
# Source: GBPIET ordinances (M.Tech) and Academic Council minutes 2025
# (B.Tech & MCA, Section 22.10).
# ---------------------------------------------------------------------------

_GRADE_TABLE_BTECH_MCA_2025 = [
    ("Letter Grade", "Grade Point", "Marks Range"),
    ("O", "10", "≥ 95%"),
    ("A+", "9.5", "90% – < 95%"),
    ("A", "9", "85% – < 90%"),
    ("B+", "8.5", "80% – < 85%"),
    ("B", "8", "75% – < 80%"),
    ("C+", "7.5", "70% – < 75%"),
    ("C", "7", "65% – < 70%"),
    ("D+", "6.5", "60% – < 65%"),
    ("D", "6", "55% – < 60%"),
    ("E+", "5.5", "50% – < 55%"),
    ("E", "5", "45% – < 50%"),
    ("P", "4.5", "40% – < 45%"),
    ("F", "0", "< 40%"),
]

_GRADE_TABLE_MTECH = [
    ("Letter Grade", "Grade Point", "Marks Range"),
    ("O — Outstanding", "10", "≥ 90%"),
    ("A+ — Excellent", "9", "85% – < 90%"),
    ("A — Very Good", "8", "80% – < 85%"),
    ("B — Good", "7", "70% – < 80%"),
    ("C — Average", "6", "60% – < 70%"),
    ("P — Pass", "5", "50% – < 60%"),
    ("F — Fail", "0", "< 50%"),
    ("AB — Absent", "—", "Absent"),
]


def _grade_table_for_program(program: str) -> List[Tuple[str, str, str]]:
    if program_short(program) == "M. TECH":
        return _GRADE_TABLE_MTECH
    return _GRADE_TABLE_BTECH_MCA_2025


def _grade_reference_block(program: str) -> Table:
    rows = _grade_table_for_program(program)
    t = Table(rows, colWidths=[40 * mm, 25 * mm, 50 * mm], hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e1b4b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    return t


def _sum_marks(records_subjects: List[dict], key: str) -> Tuple[int, int]:
    """Sum 'a/b' style marks columns. Returns (sum_a, sum_b)."""
    sa = sb = 0
    for s in records_subjects:
        v = (s.get(key) or "").strip()
        m = re.match(r"^(-?\d+)\s*/\s*(\d+)$", v)
        if m:
            sa += int(m.group(1)) if m.group(1) != "-" else 0
            sb += int(m.group(2))
    return sa, sb


def _compute_totals(subjects: List[dict]) -> dict:
    tc = 0
    tgp = 0.0
    for s in subjects:
        try:
            c = int(s.get("credits") or 0)
        except Exception:
            c = 0
        try:
            gp = float(s.get("grade_points") or 0)
        except Exception:
            gp = 0.0
        tc += c
        tgp += c * gp
    ext_a, ext_b = _sum_marks(subjects, "external")
    ses_a, ses_b = _sum_marks(subjects, "sessional")
    tot_a, tot_b = _sum_marks(subjects, "total")
    return {
        "credits": tc,
        "grade_points": tgp,
        "external": f"{ext_a}/{ext_b}" if ext_b else "",
        "sessional": f"{ses_a}/{ses_b}" if ses_b else "",
        "total": f"{tot_a}/{tot_b}" if tot_b else "",
    }

